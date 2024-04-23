import openpyxl

file = openpyxl.load_workbook("product.xlsx")
product_list = file["Sheet1"]
product_name = {}
total_revenue_per_product = {}
sort_price = {}
product_num = {}
totals = {}

for product_row in range(2,product_list.max_row):
    product = product_list.cell(product_row,4).value
    total_product = product_list.cell(product_row,2).value # type: ignore
    price_product = product_list.cell(product_row,3).value # type: ignore
    num_product = product_list.cell(product_row,1).value # type: ignore
    
    if product is not None:
        if product in product_name:
            current_num_price = product_name[product]
            product_name[product] = current_num_price +1
            
        else:
            product_name[product] = 1
    # mengumpulkan jumlah dari tiap produk
    if total_product is not None:
        total = float(total_product)     # type: ignore
        if product in totals :
            current_total_revenue = totals[product]
            totals[product] = current_total_revenue + total
        else:
            totals[product] = total 
    # mengumpulkan total revenue per product
    if (total_product is not None) and  (price_product is not None):
        total = float(total_product)     # type: ignore
        price = float(price_product) # type: ignore
        if product in total_revenue_per_product :
            current_total_revenue = total_revenue_per_product[product]
            total_revenue_per_product[product] = current_total_revenue + (total * price)
        else:
            total_revenue_per_product[product] = total * price
    # mengumpulkan price product
    if price_product is not None:
        price = float(price_product) # type: ignore
        if product in sort_price :
            current_total_revenue = sort_price[product]
            tmp = price
            if current_total_revenue > tmp :
                sort_price[product] = current_total_revenue
            else:
                sort_price[product] = tmp
        else:
            sort_price[product] = price
    # mengumpulkan num_product
    if num_product is not None:
        num = int(num_product) # type: ignore
        if product in product_num :
            current_num_product = product_num[product]
            tmp = num
            if current_num_product > tmp :
                product_num[product] = current_total_revenue
            else:
                product_num[product] = tmp
        else:
            product_num[product] = num
# Sort total_revenue_per_product by price (descending order)
sorted_prices_desc = sorted(sort_price.items(), key=lambda x: x[1], reverse=True)[:10]

# Print sorted prices in descending order
print("Top 10 highest price\n")
for product, price in sorted_prices_desc:
    print(f"{product}: {price}")

# Sort total_revenue_per_product by price (ascending order)
sorted_prices_desc = sorted(sort_price.items(), key=lambda x: x[1])[:10]

# Print sorted prices in descending order
print("\nTop 10 Lowest price\n")
for product, price in sorted_prices_desc:
    print(f"{product}: {price}")

#Keyboard Total Price
value_keyboard = total_revenue_per_product.get('Keyboard')
if value_keyboard is not None :
    print("Total price all of keyboard is",value_keyboard)

# Find the product with the lowest total product
lowest_product = min(totals, key=totals.get) # type: ignore

# Get the product number for the lowest product
lowest_product_number = product_num.get(lowest_product)

print("\nProduct with the lowest total product:", lowest_product)
print("Product number id for the lowest product:", lowest_product_number)

#Sum all revenue product and save to new file
sum_revenue = sum(total_revenue_per_product.values())
print(f"\nSum all product Revenue: {sum_revenue}")
product_list.cell(row=51,column=1,value="Sum Revenue")
product_list.cell(row=51,column=2,value="")
product_list.cell(row=51,column=3,value= sum_revenue) # type: ignore
product_list.cell(row=51,column=4,value="")
file.save("product_updt.xlsx")
file.close()